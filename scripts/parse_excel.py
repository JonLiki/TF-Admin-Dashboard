"""
parse_excel.py
--------------
Parses Block 1 Excel files (Weigh-In, KM, Lifestyle, Attendance) and
produces prisma/data/imported_data.json ready for seed.ts to consume.

Output JSON format:
[
  {
    "firstName": "Misty",
    "lastName": "Fifita",
    "team": "Group 1",
    "weights": [108.6, 108.6, 107.1, 107.1, 106.6, 107.9, 107.6, 107.2],  // Week 1..8 (null if missing)
    "startWeight": 108.6,
    "km": [0.0, 10.4, 16.6, 9.9, 11.2, 5.6, 8.7, 0.0],                    // Week 1..8
    "lifestyle": [1, 5, 5, 2, 5, 4, 5, 0],                                  // Week 1..8
    "attendance": {                                                           // session date -> present bool
      "2026-01-19": true,
      "2026-01-21": true,
      ...
    }
  },
  ...
]
"""

import openpyxl
import json
import os
from datetime import datetime, timedelta

DOCS_DIR = r'C:\Users\sione.likiliki\Documents\TF-Dashboard\docs'
OUTPUT_PATH = r'C:\Users\sione.likiliki\Documents\TF-Dashboard\prisma\data\imported_data.json'

WEIGHIN_FILE    = os.path.join(DOCS_DIR, 'Block 1 Weigh-In.xlsx')
KM_FILE         = os.path.join(DOCS_DIR, 'Block 1 KM.xlsx')
LIFESTYLE_FILE  = os.path.join(DOCS_DIR, 'Block 1 Lifestyle.xlsx')
ATTENDANCE_FILE = os.path.join(DOCS_DIR, 'Block 1 Attendance.xlsx')

def member_key(first_name: str, last_name: str) -> str:
    """Normalised key to match members across files."""
    return f"{first_name.strip().lower()}|{last_name.strip().lower()}"

def parse_float(val) -> float | None:
    """Return float or None for missing/dash values."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ('-', '', 'None'):
        return None
    try:
        return float(s.replace(',', '.'))
    except ValueError:
        return None

def parse_int(val) -> int | None:
    """Return int or None for missing/dash values."""
    f = parse_float(val)
    return int(f) if f is not None else None

# ────────────────────────────────────────────────
# 1. Parse Weigh-In
# ────────────────────────────────────────────────
def parse_weighin():
    """
    Columns: First Name, Last Name, Team, Start Weight, Current Weight,
             Overall Change, Week 1..Week 8
    Returns dict: member_key -> {firstName, lastName, team, startWeight, weights[8]}
    """
    wb = openpyxl.load_workbook(WEIGHIN_FILE)
    ws = wb.active

    members = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        first, last, team = row[0], row[1], row[2]
        if not first or str(first).strip().upper() == 'TEAM TOTAL':
            continue

        start_weight = parse_float(row[3])
        # Columns 6..14 are Week 1..Week 9 (index 6-14)
        week_weights = [parse_float(row[6 + w]) for w in range(9)]

        key = member_key(first, last or '')
        members[key] = {
            'firstName': str(first).strip(),
            'lastName': str(last).strip() if last else '',
            'team': str(team).strip() if team else '',
            'startWeight': start_weight,
            'weights': week_weights,
        }
    print(f"  Weigh-In: parsed {len(members)} members")
    return members

# ────────────────────────────────────────────────
# 2. Parse KM
# ────────────────────────────────────────────────
def parse_km():
    """
    Columns: First Name, Last Name, Team, Total KM, Week 1..Week 8 (Wait, does KM have Week 9 yet?)
    Checking inspect_output.txt: KM has Week 1..Week 8 only.
    Returns dict: member_key -> km[8 or 9]
    """
    wb = openpyxl.load_workbook(KM_FILE)
    ws = wb.active

    members = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        first, last = row[0], row[1]
        if not first or str(first).strip().upper() == 'TEAM TOTAL':
            continue
        # We'll read as many weeks as are actually in the file, but target 9 if possible.
        # Based on inspection, it's 8. We'll fill the rest with 0.
        week_km = [parse_float(row[4 + w]) or 0.0 for w in range(len(row) - 4)]
        if len(week_km) < 9:
            week_km.extend([0.0] * (9 - len(week_km)))
        
        key = member_key(first, last or '')
        members[key] = week_km
    print(f"  KM:       parsed {len(members)} members")
    return members

# ────────────────────────────────────────────────
# 3. Parse Lifestyle
# ────────────────────────────────────────────────
def parse_lifestyle():
    """
    Columns: First Name, Last Name, Team, Total Posts, Week 1..Week 8
    Returns dict: member_key -> lifestyle[8 or 9]
    """
    wb = openpyxl.load_workbook(LIFESTYLE_FILE)
    ws = wb.active

    members = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        first, last = row[0], row[1]
        if not first or str(first).strip().upper() == 'TEAM TOTAL':
            continue
        week_lifestyle = [parse_int(row[4 + w]) or 0 for w in range(len(row) - 4)]
        if len(week_lifestyle) < 9:
            week_lifestyle.extend([0] * (9 - len(week_lifestyle)))

        key = member_key(first, last or '')
        members[key] = week_lifestyle
    print(f"  Lifestyle: parsed {len(members)} members")
    return members

# ────────────────────────────────────────────────
# 4. Parse Attendance
# ────────────────────────────────────────────────
def parse_attendance():
    """
    Columns: First Name, Last Name, Team, Total Present, [session dates col 4..27]
    Header row contains date strings like 'Jan 19 (Monday)', 'Jan 21 (Wednesday)', etc.

    Returns dict: member_key -> {session_date_iso_str: bool}
    Also returns list of session date ISO strings in order.
    """
    wb = openpyxl.load_workbook(ATTENDANCE_FILE)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    # Parse session dates from header (columns 4 onwards)
    session_dates = []
    for col_val in header[4:]:
        if col_val is None:
            session_dates.append(None)
            continue
        # Format: 'Jan 19 (Monday)' — we parse the date part
        date_part = str(col_val).split('(')[0].strip()  # e.g. 'Jan 19'
        
        # Specific fix for known typo in Attendance header: 'Mar 4 (Monday)' should be 'Mar 2'
        if date_part == 'Mar 4' and '(Monday)' in str(col_val):
            date_part = 'Mar 2'

        try:
            # Assume year 2026 (Block 1)
            dt = datetime.strptime(f"{date_part} 2026", "%b %d %Y")
            session_dates.append(dt.strftime('%Y-%m-%d'))
        except ValueError:
            print(f"  WARNING: Could not parse date header: '{col_val}'")
            session_dates.append(None)

    members = {}
    for i, row in enumerate(rows[1:], start=1):
        first, last = row[0], row[1]
        if not first or str(first).strip().upper() == 'TEAM TOTAL':
            continue

        attendance_map = {}
        for j, date_str in enumerate(session_dates):
            if date_str is None:
                continue
            cell_val = row[4 + j] if (4 + j) < len(row) else None
            is_present = str(cell_val).strip().lower() == 'present' if cell_val is not None else False
            attendance_map[date_str] = is_present

        key = member_key(first, last or '')
        members[key] = attendance_map

    print(f"  Attendance: parsed {len(members)} members across {len([d for d in session_dates if d])} sessions")
    return members, [d for d in session_dates if d]

# ────────────────────────────────────────────────
# 5. Merge and Output
# ────────────────────────────────────────────────
def main():
    print("Parsing Excel files...")
    weighin_data = parse_weighin()
    km_data      = parse_km()
    life_data    = parse_lifestyle()
    att_data, session_dates = parse_attendance()

    # Merge all data by member key (weight data is the source of truth for members)
    output = []
    for key, wdata in weighin_data.items():
        record = {
            'firstName':   wdata['firstName'],
            'lastName':    wdata['lastName'],
            'team':        wdata['team'],
            'startWeight': wdata['startWeight'],
            'weights':     wdata['weights'],   # list of 8 (float or null)
            'km':          km_data.get(key, [0.0] * 8),
            'lifestyle':   life_data.get(key, [0] * 8),
            'attendance':  att_data.get(key, {}),
        }
        output.append(record)

    # Warn about members in attendance/km/lifestyle but not weighin
    all_keys = set(km_data) | set(life_data) | set(att_data)
    missing = all_keys - set(weighin_data)
    if missing:
        print(f"  WARNING: These members appear in KM/Lifestyle/Attendance but NOT in Weigh-In: {missing}")

    # Ensure output dir exists
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump({
            'members': output,
            'sessionDates': session_dates,
        }, f, indent=2, ensure_ascii=False)

    print(f"\nDone! Wrote {len(output)} members to:")
    print(f"  {OUTPUT_PATH}")
    print(f"\nSample record:")
    if output:
        sample = output[0]
        print(f"  {sample['firstName']} {sample['lastName']} ({sample['team']})")
        print(f"  startWeight: {sample['startWeight']}")
        print(f"  weights (W1-W9): {sample['weights']}")
        print(f"  km      (W1-W9): {sample['km']}")
        print(f"  lifestyle(W1-W9): {sample['lifestyle']}")
        att_summary = {k: v for k, v in list(sample['attendance'].items())[:4]}
        print(f"  attendance (first 4 sessions): {att_summary}")

if __name__ == '__main__':
    main()
